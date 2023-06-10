"""
读取数据库（暂时是excel）并统计省份->国家信息
"""
# 导入requests包
import json
import re

import openpyxl
import pycountry
import requests
from countryinfo import CountryInfo
from tqdm import tqdm

# 替换为您的 ArcGIS REST API 地址
from geopy import Nominatim

geocode_url = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates"
reverse_geocode_url = "https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/reverseGeocode"

# 获得国家之间的映射
try:
    with open("./static/resources/word.json", 'r', encoding='utf-8') as file:
        # 路径和启动文件所在有关，所以这里是./而不是../
        countryMap = json.loads(file.read())
        en2chn = countryMap['namemap']
except:
    with open("../static/resources/word.json", 'r', encoding='utf-8') as file:
        # 路径和启动文件所在有关，所以这里是./而不是../
        countryMap = json.loads(file.read())
        en2chn = countryMap['namemap']


def geocode(location):
    params = {
        "SingleLine": location,
        "outFields": "Country",
        "f": "json"
    }
    response = requests.get(geocode_url, params=params).json()

    if "candidates" in response and len(response["candidates"]) > 0:
        predict_ = response["candidates"][0]
        country = predict_["attributes"]["Country"]
        location_ = (predict_['location']['x'], predict_['location']['y'])
        return country, location_

    return None


def abbrev2name(country_code):
    import pycountry
    try:
        country = pycountry.countries.get(alpha_2=country_code)
        if country:
            country_name = country.name
            print(f"{country_code}: {country_name}")
            return country_name
        else:
            return "找不到该国家缩写的全称。"
    except Exception as e:
        return str(e)


# 根据国家缩写获取国家全称
def get_country_name(country_code):
    try:
        country = CountryInfo(country_code)
        country_name = country.info()['name']
        return country_name
    except Exception as e:
        return "找不到该国家缩写的全称。"


def reverGeo_gaode(province_str=""):
    url = "https://restapi.amap.com/v3/geocode/geo"
    parameters = {"address": province_str,
                  "output": "JSON",
                  "key": "c93f574233412333a6b3779b2fe789e7"}
    res = eval(requests.get(url=url, params=parameters).text)
    if res.get("status") is "0":
        print("查询无果")
        return None
    print(res.get("geocodes")[0])
    print(res.get("geocodes")[0].get("country"))


def reverse_geocode(lat, lon):
    params = {
        "location": f"{lon},{lat}",
        "outFields": "Country",
        "f": "json"
    }
    response = requests.get(reverse_geocode_url, params=params).json()

    if "address" in response and "CountryCode" in response["address"]:
        country = response["address"]["CountryCode"]
        return country

    return None


def get_map_data():
    """
    dataArr
    [
    {
      "name": "阿富汗",
      "value": 28397.812
    },
    :return:
    """
    # 读取文件内容
    try:
        origin_book_path = "./static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    except:
        origin_book_path = "../static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    sheet_instance = work_book["Sheet1"]
    positions = sheet_instance['C'][1:]
    # 获取相应信息
    val_dict = {}
    locations = []
    dataArr = []
    for position in tqdm(positions):
        try:
            if position.value is None:
                continue
            country_, location_ = geocode(position.value)
            locations.append(location_)
            try:
                country_name = pycountry.countries.get(alpha_2=country_).name
                val_dict[country_name] = val_dict.setdefault(country_name, 0) + 1
            except:
                try:
                    country_name = pycountry.countries.get(alpha_3=country_).name
                    val_dict[country_name] = val_dict.setdefault(country_name, 0) + 1
                except:
                    pass
        except:
            pass
    for key in val_dict.keys():
        dataArr.append({"name": key, "value": val_dict.get(key)})
    with open('./dataArr.txt', 'w', encoding='utf-8') as writer:
        writer.write(str(locations))
    with open('./locations_.txt', 'w', encoding='utf-8') as writer:
        writer.write(str(locations))
    dataArrchn_ = en2chn()
    return dataArrchn_, locations


def get_locations():
    try:
        origin_book_path = "./static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    except:
        origin_book_path = "../static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    sheet_instance = work_book["Sheet1"]
    positions = sheet_instance['C'][1:]
    # 获取相应信息
    locations = []
    for position in tqdm(positions):
        try:
            if position.value is None:
                continue
            country_, location_ = geocode(position.value)
            locations.append({"name": position, "coord": list(location_)})
        except:
            pass
    with open('./locations_list.txt', 'w', encoding='utf-8') as writer:
        writer.write(str(locations))


def again():
    try:
        origin_book_path = "./static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    except:
        origin_book_path = "../static/resources/docs/userInfo.xlsx"
        work_book = openpyxl.load_workbook(origin_book_path)
    sheet_instance = work_book["Sheet1"]
    positions_raw = sheet_instance['C']  # 索引从0开始，所以刚好纠正过来
    positions_val_list = [position.value for position in positions_raw]
    with open('./locations_list_.txt', 'r', encoding='utf-8') as reader:
        locations_str = reader.read()

    def replace(match):
        return match.group(1)

    # 使用正则表达式进行替换
    result = re.sub(r"<.*?\.C(.*?)>", replace, locations_str)
    locations = eval(result)
    for location in locations:
        location["name"] = positions_val_list[location["name"] - 1].replace("'", "")

    with open('./locations_val_coord_single.txt', 'w', encoding='utf-8') as writer:
        writer.write(str(locations))


def read_cache():
    try:
        with open('./dataArr_chn.txt', 'r', encoding='utf-8') as reader:
            dataArr = eval(reader.read())
        with open('./locations_val_coord_single.txt', 'r', encoding='utf-8') as reader:
            locations = eval(reader.read())
    except:
        with open('./utils/dataArr_chn.txt', 'r', encoding='utf-8') as reader:
            dataArr = eval(reader.read())
        print(len(dataArr))
        with open('./utils/locations_val_coord_single.txt', 'r', encoding='utf-8') as reader:
            locations = eval(reader.read())
    return dataArr, locations


def eng2chn():
    with open('./dataArr.txt', 'r', encoding='utf-8') as reader:
        dataArr = eval(reader.read())
    dataArr_ = []
    for dict_ in dataArr:
        dataArr_.append({"name": en2chn.get(dict_.get("name")), "value": dict_.get("value")})
    with open('./dataArr_chn.txt', 'w', encoding='utf-8') as writer:
        writer.write(str(dataArr_))
    return dataArr_


if __name__ == '__main__':
    again()
